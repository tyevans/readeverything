"""Spreadsheets, OOXML and ODF.

`CellRange`'s only producer. A sheet is addressed as cells, not as a character
offset into whatever delimiter this module happens to join them with — that is
the whole argument for the locator, and it is why the outline and every
affordance cite cells rather than spans.

**Formulas and values are both reachable, and that is the point.** `represent`
shows the VALUE, because that is what the sheet means to a reader.
`read_cells(..., formulas=True)` shows the formula, because that is what an
auditor needs. Reporting only one of the two is how a spreadsheet lies.

A third case exists that neither of those covers, and it is the one that bites:
a formula whose cached value was never written. openpyxl computes nothing, so
every workbook it produced is in this state, and plenty of other writers are
too. Rendering those cells blank would report "this sheet is empty" about a
sheet full of arithmetic — this library's signature failure. So the formula
text is shown in their place and a `Degradation` names both the cause and the
way out.

`describe` loads with `read_only=True` and closes: a million-row workbook must
not be materialised to answer "what sheets are there".

openpyxl is imported directly here, guarded exactly as `pdf.py` guards
pypdfium2.
"""

from __future__ import annotations

import io
import re
import time
from dataclasses import dataclass
from typing import ClassVar

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - exercised via a patched sys.modules
    raise ImportError(
        "readeverything's spreadsheet support needs openpyxl, which ships in the "
        "'office' extra: pip install 'readeverything[office]'. "
        "The composition root omits spreadsheet handling when openpyxl is absent, so "
        "reaching this means the handler was imported directly."
    ) from exc
from pydantic import BaseModel, Field

from readeverything.adapters.odf import odf_sheets
from readeverything.adapters.ooxml import ODF_SHEETS_MIME, SHEETS_MIME, office_mimetype
from readeverything.domain.affordance import Affordance, DetailLevel
from readeverything.domain.capability import Capability
from readeverything.domain.card import Card, Segment
from readeverything.domain.errors import UnknownAffordanceError
from readeverything.domain.identity import MediaKind, SourceRef
from readeverything.domain.locator_map import LocatorMap, LocatorSegment
from readeverything.domain.locators import ByteRange, CellRange, CharSpan
from readeverything.domain.observation import OperationFinished, OperationStarted
from readeverything.domain.rendition import (
    Budget,
    Degradation,
    Rendered,
    Rendition,
    TextContent,
)
from readeverything.ports.observation import Observer, emit
from readeverything.ports.source import SourceReader

#: What `represent` calls itself when it narrates, matching every other handler.
_OPERATION = "represent"

#: How a row's cells are joined. Matches the Word handler's table rendering, so
#: tabular text reads the same whichever family it came out of.
CELL_DELIMITER = " | "

#: Every sheet's text ends with this, and the sheet's `LocatorSegment` INCLUDES
#: it. `LocatorMap` demands total, gapless, zero-start coverage and
#: `CharSpan.__post_init__` rejects `start >= end`, so an empty sheet would
#: otherwise contribute a zero-width span and raise. Owning the separator means
#: every sheet owns at least one character no matter what it contained. Do not
#: "simplify" this away: an empty sheet between two full ones breaks the map.
SHEET_SEPARATOR = "\n"

#: What stands in the flattened text for a sheet with no used cells. A sheet
#: that exists and is empty is a different fact from a sheet that is missing.
_EMPTY_SHEET = "(this sheet has no cells)"

#: How many cells are named individually in the uncomputed-formula degradation
#: before it summarises the rest. One report per workbook, not one per cell,
#: matching how `pdf.py` reports scanned pages.
_LISTED_CELLS = 10

#: `A1`, `AA12`, or a `A1:C3` block. Anchored, so a string with a range buried
#: inside it does not parse as one — a caller who typed something else gets a
#: degradation rather than a citation to a cell they never named.
_A1_CELL = re.compile(r"^([A-Za-z]+)([1-9][0-9]*)$")

#: Base-26 with no zero digit: A is 0, Z is 25, AA is 26.
_ALPHABET_SIZE = 26


def to_a1(row: int, col: int) -> str:
    """A 0-indexed `(row, col)` as A1 notation.

    Presentation, deliberately living in the handler rather than in
    `domain/locators.py`: A1 is 1-indexed and its columns are base-26 letters,
    neither of which the domain has any business knowing.
    """
    letters = ""
    remaining = col
    while True:
        letters = chr(ord("A") + remaining % _ALPHABET_SIZE) + letters
        remaining = remaining // _ALPHABET_SIZE - 1
        if remaining < 0:
            break
    return f"{letters}{row + 1}"


def _cell(reference: str) -> tuple[int, int] | None:
    match = _A1_CELL.match(reference.strip())
    if match is None:
        return None
    letters, digits = match.groups()
    col = 0
    for letter in letters.upper():
        col = col * _ALPHABET_SIZE + (ord(letter) - ord("A") + 1)
    return int(digits) - 1, col - 1


def parse_a1(a1_range: str, sheet: str) -> CellRange | None:
    """An A1 cell or block as a `CellRange`, or None if it does not parse.

    None rather than an exception, and None rather than a guess: a caller who
    typed something this cannot read gets a degradation naming what it tried,
    instead of a citation to a cell nobody asked for.

    A reversed range (`C3:A1`) names the same block as `A1:C3`, so the corners
    are ordered before the extent is computed — otherwise the extent is
    negative and `CellRange` raises on input the caller was entitled to send.
    """
    parts = a1_range.split(":")
    if len(parts) == 1:
        cell = _cell(parts[0])
        if cell is None:
            return None
        return CellRange(sheet=sheet, row=cell[0], col=cell[1])
    if len(parts) != 2:
        return None
    first, second = _cell(parts[0]), _cell(parts[1])
    if first is None or second is None:
        return None
    top, left = min(first[0], second[0]), min(first[1], second[1])
    bottom, right = max(first[0], second[0]), max(first[1], second[1])
    return CellRange(sheet=sheet, row=top, col=left, rows=bottom - top + 1, cols=right - left + 1)


def _render(value: object) -> str:
    """One cell as text. `None` is an empty cell and renders as nothing."""
    return "" if value is None else str(value)


@dataclass(frozen=True, slots=True)
class _Sheet:
    """One sheet's name and its rows of already-rendered cell text."""

    name: str
    rows: tuple[tuple[str, ...], ...]

    @property
    def columns(self) -> int:
        return max((len(row) for row in self.rows), default=0)

    def rendered(self) -> str:
        if not self.rows:
            return f"{self.name}\n{_EMPTY_SHEET}"
        body = "\n".join(CELL_DELIMITER.join(row) for row in self.rows)
        return f"{self.name}\n{body}"

    def locator(self) -> CellRange:
        return CellRange(
            sheet=self.name,
            row=0,
            col=0,
            rows=max(1, len(self.rows)),
            cols=max(1, self.columns),
        )


@dataclass(frozen=True, slots=True)
class _Workbook:
    """Every sheet, plus the cells whose formulas were never computed."""

    sheets: tuple[_Sheet, ...]
    uncomputed: tuple[str, ...]


class ReadSheetParams(BaseModel):
    name: str = Field(default="", description="Sheet name. Empty means the first sheet.")
    offset: int = Field(default=0, ge=0, description="0-indexed first row to return.")
    limit: int = Field(default=100, gt=0, description="How many rows to return.")


class ReadCellsParams(BaseModel):
    name: str = Field(default="", description="Sheet name. Empty means the first sheet.")
    a1_range: str = Field(default="A1", description="A cell or block in A1 notation, e.g. B2:D10.")
    formulas: bool = Field(
        default=False,
        description="Show each cell's formula instead of its value. Use this to audit a sheet.",
    )


class ListSheetsParams(BaseModel):
    pass


class OfficeSheetsHandler:
    """Reads a workbook, and maps every character to the sheet it came from."""

    mime_patterns: ClassVar[tuple[str, ...]] = (SHEETS_MIME, ODF_SHEETS_MIME)
    priority: ClassVar[int] = 0
    handler_id: ClassVar[str] = "office_sheets"
    handler_version: ClassVar[int] = 1

    def __init__(self, *, source: SourceReader, observer: Observer | None = None) -> None:
        self._source = source
        self._observer = observer

    def requires(self) -> frozenset[Capability]:
        """Nothing. Reading a workbook needs no model and no binary."""
        return frozenset()

    def affordances(self) -> tuple[Affordance, ...]:
        return (
            Affordance(
                name="read_sheet",
                description="Return a page of one sheet's rows as delimited text.",
                params=ReadSheetParams,
                requires=frozenset(),
                level=DetailLevel.SEGMENT,
            ),
            Affordance(
                name="read_cells",
                description=(
                    "Return one cell or block in A1 notation. Pass formulas=true to see "
                    "each cell's formula instead of its value."
                ),
                params=ReadCellsParams,
                requires=frozenset(),
                level=DetailLevel.SEGMENT,
            ),
            Affordance(
                name="list_sheets",
                description="List every sheet with its used range.",
                params=ListSheetsParams,
                requires=frozenset(),
                level=DetailLevel.SEGMENT,
            ),
        )

    # -- parsing ---------------------------------------------------------

    def _is_odf(self, data: bytes) -> bool:
        return office_mimetype(data) == ODF_SHEETS_MIME

    def _from_odf(self, data: bytes) -> _Workbook | None:
        sheets = odf_sheets(data)
        if not sheets:
            return None
        # ODF cells arrive already flattened to text through this reader, with
        # no formula/value distinction to lose — so the uncomputed-formula
        # report never fires for an `.ods`, which is honest: nothing was hidden.
        return _Workbook(
            sheets=tuple(_Sheet(name=name, rows=tuple(rows)) for name, rows in sheets),
            uncomputed=(),
        )

    def _from_ooxml(self, data: bytes) -> _Workbook | None:
        """Both loads, because one of them cannot tell the interesting case.

        `data_only=True` gives cached values and `None` for a formula that was
        never computed; the plain load gives the formula text. Only holding both
        distinguishes an empty cell from an uncomputed one, and that distinction
        is what stops this handler reporting a sheet of arithmetic as empty. Two
        opens of an in-memory buffer are cheap; guessing is not.
        """
        try:
            values = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            formulas = openpyxl.load_workbook(io.BytesIO(data))
        except Exception:
            return None
        sheets: list[_Sheet] = []
        uncomputed: list[str] = []
        try:
            for name in values.sheetnames:
                value_sheet = values[name]
                formula_sheet = formulas[name]
                rows: list[tuple[str, ...]] = []
                for value_row, formula_row in zip(
                    value_sheet.iter_rows(), formula_sheet.iter_rows(), strict=False
                ):
                    cells: list[str] = []
                    for value_cell, formula_cell in zip(value_row, formula_row, strict=False):
                        formula = formula_cell.value
                        is_formula = isinstance(formula, str) and formula.startswith("=")
                        if is_formula and value_cell.value is None:
                            uncomputed.append(
                                f"{name}!{to_a1(value_cell.row - 1, value_cell.column - 1)}"
                            )
                            cells.append(_render(formula))
                        else:
                            cells.append(_render(value_cell.value))
                    rows.append(tuple(cells))
                sheets.append(_Sheet(name=name, rows=tuple(rows)))
        except Exception:
            return None
        return _Workbook(sheets=tuple(sheets), uncomputed=tuple(uncomputed))

    def _parse(self, data: bytes) -> _Workbook | None:
        if self._is_odf(data):
            return self._from_odf(data)
        return self._from_ooxml(data)

    def _formula_sheets(self, data: bytes) -> tuple[_Sheet, ...] | None:
        """Every sheet with formulas shown in place of values, for auditing."""
        if self._is_odf(data):
            parsed = self._from_odf(data)
            return None if parsed is None else parsed.sheets
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(data))
            return tuple(
                _Sheet(
                    name=name,
                    rows=tuple(
                        tuple(_render(cell.value) for cell in row)
                        for row in workbook[name].iter_rows()
                    ),
                )
                for name in workbook.sheetnames
            )
        except Exception:
            return None

    def _pick(self, sheets: tuple[_Sheet, ...], name: str) -> _Sheet | None:
        if not name:
            return sheets[0] if sheets else None
        return next((sheet for sheet in sheets if sheet.name == name), None)

    # -- flattening ------------------------------------------------------

    def _flatten(
        self, sheets: tuple[_Sheet, ...]
    ) -> tuple[str, tuple[LocatorSegment, ...], tuple[int, ...]]:
        chunks: list[str] = []
        segments: list[LocatorSegment] = []
        barriers: list[int] = []
        cursor = 0
        for index, sheet in enumerate(sheets):
            chunk = sheet.rendered() + SHEET_SEPARATOR
            if index:
                barriers.append(cursor)
            segments.append(LocatorSegment(CharSpan(cursor, cursor + len(chunk)), sheet.locator()))
            cursor += len(chunk)
            chunks.append(chunk)
        return "".join(chunks), tuple(segments), tuple(barriers)

    def _uncomputed_degradation(self, cells: tuple[str, ...]) -> tuple[Degradation, ...]:
        """One report per workbook, not one per cell.

        The detail names the CAUSE and the WAY OUT, because a caller who only
        learns that something was degraded will conclude the sheet is empty —
        which is exactly the false claim this report exists to prevent.
        """
        if not cells:
            return ()
        head = ", ".join(cells[:_LISTED_CELLS])
        listed = (
            head if len(cells) <= _LISTED_CELLS else f"{head} and {len(cells) - _LISTED_CELLS} more"
        )
        return (
            Degradation(
                what="formulas without cached values",
                detail=(
                    f"{len(cells)} cell(s) hold a formula whose cached value was never "
                    f"written ({listed}); the workbook was saved by a tool that does not "
                    "compute formulas, so the formula text is shown in place of a result. "
                    "These cells are not empty — call read_cells with formulas=true to "
                    "audit them."
                ),
            ),
        )

    # -- the handler surface ---------------------------------------------

    async def describe(self, ref: SourceRef) -> Card:
        """Sheet names and shapes, read without materialising the cells.

        `read_only=True` is what keeps this cheap: a million-row workbook must
        not be fully loaded to answer "what sheets are there". A
        `ReadOnlyWorksheet` has no `.dimensions`, so the used range comes from
        `calculate_dimension()`.

        `kind` is `BINARY`, matching `pdf.py`: these mimetypes reach this
        handler at the registry's exact-mimetype step, long before the kind step.
        """
        data = await self._source.read_bytes(ref.uri)
        shapes = self._shapes(data)
        if shapes is None:
            return Card(
                ref=ref,
                kind=MediaKind.BINARY,
                facts={"readable": "no", "size_bytes": ref.size_bytes},
                outline=(),
                excerpt=None,
                affordances=self.affordances(),
            )
        facts: dict[str, str | int | float] = {
            "readable": "yes",
            "sheet_count": len(shapes),
            "size_bytes": ref.size_bytes,
        }
        outline: list[Segment] = []
        for name, rows, columns in shapes:
            facts[f"sheet.{name}.used_range"] = f"A1:{to_a1(max(0, rows - 1), max(0, columns - 1))}"
            facts[f"sheet.{name}.rows"] = rows
            facts[f"sheet.{name}.columns"] = columns
            outline.append(
                Segment(
                    CellRange(sheet=name, row=0, col=0, rows=max(1, rows), cols=max(1, columns)),
                    name,
                )
            )
        return Card(
            ref=ref,
            kind=MediaKind.BINARY,
            facts=facts,
            outline=tuple(outline),
            excerpt=None,
            affordances=self.affordances(),
        )

    def _shapes(self, data: bytes) -> list[tuple[str, int, int]] | None:
        """`(name, rows, columns)` per sheet, without loading the cells."""
        if self._is_odf(data):
            parsed = self._from_odf(data)
            if parsed is None:
                return None
            return [(s.name, len(s.rows), s.columns) for s in parsed.sheets]
        workbook = None
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
            return [
                (name, workbook[name].max_row or 0, workbook[name].max_column or 0)
                for name in workbook.sheetnames
            ]
        except Exception:
            return None
        finally:
            if workbook is not None:
                # A read-only workbook holds an open zip handle until closed.
                workbook.close()

    async def invoke(self, ref: SourceRef, name: str, params: BaseModel) -> Rendition:
        match name:
            case "read_sheet":
                if not isinstance(params, ReadSheetParams):
                    raise TypeError(f"expected ReadSheetParams, got {type(params).__name__}")
                return await self._read_sheet(ref, params)
            case "read_cells":
                if not isinstance(params, ReadCellsParams):
                    raise TypeError(f"expected ReadCellsParams, got {type(params).__name__}")
                return await self._read_cells(ref, params)
            case "list_sheets":
                if not isinstance(params, ListSheetsParams):
                    raise TypeError(f"expected ListSheetsParams, got {type(params).__name__}")
                return await self._list_sheets(ref)
            case _:
                raise UnknownAffordanceError(name, (a.name for a in self.affordances()))

    def _degraded(self, ref: SourceRef, detail: str) -> Rendition:
        """What every unreadable or out-of-range request returns.

        Never an exception: an agent guessing a sheet name gets a result it can
        read and correct.
        """
        return Rendition(
            locator=ByteRange(0, max(1, ref.size_bytes)),
            content=TextContent(detail),
            degraded=True,
        )

    async def _read_sheet(self, ref: SourceRef, params: ReadSheetParams) -> Rendition:
        data = await self._source.read_bytes(ref.uri)
        parsed = self._parse(data)
        if parsed is None:
            return self._degraded(ref, f"{ref.uri} could not be opened as a workbook")
        sheet = self._pick(parsed.sheets, params.name)
        if sheet is None:
            names = ", ".join(s.name for s in parsed.sheets) or "(none)"
            return self._degraded(
                ref, f"sheet {params.name!r} does not exist; the workbook has: {names}"
            )
        rows = sheet.rows[params.offset : params.offset + params.limit]
        if not rows:
            return self._degraded(
                ref,
                f"sheet {sheet.name!r} has {len(sheet.rows)} row(s); "
                f"there is nothing at offset {params.offset}",
            )
        body = "\n".join(CELL_DELIMITER.join(row) for row in rows)
        return Rendition(
            # `rows` is what came back, not what was asked for. A locator
            # claiming rows that were never read is a citation to nothing.
            locator=CellRange(
                sheet=sheet.name,
                row=params.offset,
                col=0,
                rows=len(rows),
                cols=max(1, max((len(row) for row in rows), default=1)),
            ),
            content=TextContent(body),
        )

    async def _read_cells(self, ref: SourceRef, params: ReadCellsParams) -> Rendition:
        data = await self._source.read_bytes(ref.uri)
        sheets = self._formula_sheets(data) if params.formulas else None
        if sheets is None:
            parsed = self._parse(data)
            if parsed is None:
                return self._degraded(ref, f"{ref.uri} could not be opened as a workbook")
            sheets = parsed.sheets
        sheet = self._pick(sheets, params.name)
        if sheet is None:
            names = ", ".join(s.name for s in sheets) or "(none)"
            return self._degraded(
                ref, f"sheet {params.name!r} does not exist; the workbook has: {names}"
            )
        block = parse_a1(params.a1_range, sheet.name)
        if block is None:
            return self._degraded(
                ref,
                f"{params.a1_range!r} is not a cell or range in A1 notation, "
                "which looks like B2 or B2:D10",
            )
        lines: list[str] = []
        for row in sheet.rows[block.row : block.row + block.rows]:
            lines.append(CELL_DELIMITER.join(row[block.col : block.col + block.cols]))
        return Rendition(locator=block, content=TextContent("\n".join(lines)))

    async def _list_sheets(self, ref: SourceRef) -> Rendition:
        data = await self._source.read_bytes(ref.uri)
        shapes = self._shapes(data)
        if not shapes:
            return self._degraded(ref, f"{ref.uri} could not be opened as a workbook")
        lines = [
            f"{name}: A1:{to_a1(max(0, rows - 1), max(0, columns - 1))} "
            f"({rows} row(s), {columns} column(s))"
            for name, rows, columns in shapes
        ]
        return Rendition(
            locator=ByteRange(0, max(1, ref.size_bytes)),
            content=TextContent("\n".join(lines)),
        )

    async def represent(self, ref: SourceRef, budget: Budget) -> Rendered:
        """Narrated start to finish, matching every other handler."""
        emit(self._observer, OperationStarted(operation=_OPERATION, ref=ref))
        started = time.perf_counter()
        try:
            return await self._represent(ref, budget)
        finally:
            emit(
                self._observer,
                OperationFinished(
                    operation=_OPERATION, ref=ref, elapsed_s=time.perf_counter() - started
                ),
            )

    async def _represent(self, ref: SourceRef, budget: Budget) -> Rendered:
        data = await self._source.read_bytes(ref.uri)
        parsed = self._parse(data)
        if parsed is None:
            return self._nothing_to_read(
                ref,
                budget,
                summary=f"Unreadable workbook {ref.uri}, {ref.size_bytes} bytes.",
                what="workbook unopenable",
                detail="the file could not be opened as a workbook; no cells were read",
            )
        if not parsed.sheets:
            return self._nothing_to_read(
                ref,
                budget,
                summary=f"Workbook {ref.uri} has no sheets.",
                what="workbook has no sheets",
                detail="the workbook opened but contains no sheets; no cells were read",
            )
        text, segments, barriers = self._flatten(parsed.sheets)
        return self._fit(
            text, segments, barriers, budget, self._uncomputed_degradation(parsed.uncomputed)
        )

    def _nothing_to_read(
        self, ref: SourceRef, budget: Budget, *, summary: str, what: str, detail: str
    ) -> Rendered:
        """A rendition for a file with no sheet to point at.

        Located by `ByteRange` rather than `CellRange`: no sheet was ever
        observed, and naming one would be a claim about a workbook this handler
        never opened.
        """
        segments = (
            LocatorSegment(CharSpan(0, len(summary)), ByteRange(0, max(1, ref.size_bytes))),
        )
        return self._fit(summary, segments, (), budget, (Degradation(what=what, detail=detail),))

    def _fit(
        self,
        full: str,
        segments: tuple[LocatorSegment, ...],
        barriers: tuple[int, ...],
        budget: Budget,
        degradations: tuple[Degradation, ...],
    ) -> Rendered:
        """Apply the budget, pruning the map and the barriers along with the text.

        A million-row sheet is cut here, and the cut is announced — silent
        truncation is invisible in exactly the case where the answer is wrong.
        """
        if budget.max_chars is None or len(full) <= budget.max_chars:
            return Rendered(
                text=full,
                locator_map=LocatorMap.build(segments),
                barriers=barriers,
                degradations=degradations,
            )
        keep = max(1, budget.max_chars)
        text = full[:keep]
        kept = tuple(
            LocatorSegment(CharSpan(s.span.start, min(s.span.end, keep)), s.locator)
            for s in segments
            if s.span.start < keep
        )
        return Rendered(
            text=text,
            locator_map=LocatorMap.build(kept),
            barriers=tuple(barrier for barrier in barriers if barrier < keep),
            degradations=(
                *degradations,
                Degradation(
                    what="text truncated",
                    detail=f"kept {len(text)} of {len(full)} characters",
                ),
            ),
        )
